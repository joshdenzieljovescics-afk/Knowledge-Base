"""
SafexpressOps - Workload Analysis Agent (IMPROVED)
---------------------------------------------------
Enhanced parser for VFP-Productivity_TMS.xlsx format

KEY IMPROVEMENTS:
1. Detects TIME STUDY data from actual Excel structure (Row 9: T1-T10 trials)
2. Extracts volume from SUMMARY sheet (Average Volume/Day column)
3. Handles multiple unit types per process (PALLET, BOX, ITEMS, PCS)
4. Flexible detection - works with their actual data layout
"""

from fastapi import FastAPI, HTTPException, File, UploadFile
from pydantic import BaseModel
from typing import Any, List, Dict, Optional
from datetime import datetime
import pandas as pd
import numpy as np
import io
import openpyxl
import json
import os
import time
from functools import wraps
import requests
import asyncio

MONITORING_URL = os.getenv("MONITORING_SERVICE_URL", "http://localhost:8009")

# Import workload analyzer modules
from workload_analyzer import (
    WorkloadAnalyzer,
    TimeStudyObservation,
    WorkloadParameters,
    ProcessComparator,
)

app = FastAPI(
    title="Workload Analysis Agent",
    description="Dedicated agent for workforce planning and capacity analysis",
    version="2.0.0",
)

# Global state for caching
analysis_cache = {}
analysis_history = []


# ============================================================================
# MONITORING UTILITIES
# ============================================================================


def calculate_accuracy(result: Any, task_type: str) -> float:
    """Calculate task-specific accuracy score"""
    if not isinstance(result, dict):
        return 100.0

    if result.get("status") != "success":
        return 0.0

    # Task-specific accuracy calculation
    if task_type == "workload_analysis":
        processes = result.get("summary", {}).get("total_processes", 0)
        if processes > 0:
            # Check if bottleneck was identified
            bottleneck = result.get("bottleneck", {})
            if bottleneck and bottleneck.get("process_name"):
                return 100.0
            return 80.0  # Partial success
        return 0

    else:
        return 100.0


def monitor_task(agent_name: str, task_type: str):
    """
    Decorator to monitor agent tasks (supports both sync and async)
    """

    def decorator(func):
        if asyncio.iscoroutinefunction(func):
            # Async version
            @wraps(func)
            async def async_wrapper(*args, **kwargs):
                task_id = f"{agent_name}_{int(time.time() * 1000)}"
                start_time = time.time()

                try:
                    # Execute the async function
                    result = await func(*args, **kwargs)

                    # Calculate metrics
                    latency = time.time() - start_time
                    success = (
                        result.get("status") == "success"
                        if isinstance(result, dict)
                        else True
                    )

                    # Calculate accuracy
                    accuracy = calculate_accuracy(result, task_type)

                    # Report to monitoring
                    try:
                        requests.post(
                            f"{MONITORING_URL}/metrics/record",
                            json={
                                "agent_name": agent_name,
                                "task_id": task_id,
                                "timestamp": datetime.now().isoformat(),
                                "accuracy_score": accuracy,
                                "latency_seconds": latency,
                                "success": success,
                                "error_message": (
                                    result.get("error")
                                    if isinstance(result, dict)
                                    else None
                                ),
                                "task_type": task_type,
                                "input_size": len(str(kwargs)) if kwargs else 0,
                                "output_size": len(str(result)) if result else 0,
                            },
                            timeout=2,
                        )
                        print(
                            f"   📊 Monitoring: {task_type} | Success: {success} | Accuracy: {accuracy:.1f}%"
                        )
                    except Exception as e:
                        print(f"   ⚠️ Monitoring report failed: {str(e)}")

                    return result

                except Exception as e:
                    latency = time.time() - start_time

                    # Report failure
                    try:
                        requests.post(
                            f"{MONITORING_URL}/metrics/record",
                            json={
                                "agent_name": agent_name,
                                "task_id": task_id,
                                "timestamp": datetime.now().isoformat(),
                                "accuracy_score": 0,
                                "latency_seconds": latency,
                                "success": False,
                                "error_message": str(e),
                                "task_type": task_type,
                            },
                            timeout=2,
                        )
                    except:
                        pass

                    raise

            return async_wrapper
        else:
            # Sync version
            @wraps(func)
            def sync_wrapper(*args, **kwargs):
                task_id = f"{agent_name}_{int(time.time() * 1000)}"
                start_time = time.time()

                try:
                    # Execute the function
                    result = func(*args, **kwargs)

                    # Calculate metrics
                    latency = time.time() - start_time
                    success = (
                        result.get("status") == "success"
                        if isinstance(result, dict)
                        else True
                    )

                    # Calculate accuracy
                    accuracy = calculate_accuracy(result, task_type)

                    # Report to monitoring
                    try:
                        requests.post(
                            f"{MONITORING_URL}/metrics/record",
                            json={
                                "agent_name": agent_name,
                                "task_id": task_id,
                                "timestamp": datetime.now().isoformat(),
                                "accuracy_score": accuracy,
                                "latency_seconds": latency,
                                "success": success,
                                "error_message": (
                                    result.get("error")
                                    if isinstance(result, dict)
                                    else None
                                ),
                                "task_type": task_type,
                                "input_size": len(str(kwargs)) if kwargs else 0,
                                "output_size": len(str(result)) if result else 0,
                            },
                            timeout=2,
                        )
                        print(
                            f"   📊 Monitoring: {task_type} | Success: {success} | Accuracy: {accuracy:.1f}%"
                        )
                    except Exception as e:
                        print(f"   ⚠️ Monitoring report failed: {str(e)}")

                    return result

                except Exception as e:
                    latency = time.time() - start_time

                    # Report failure
                    try:
                        requests.post(
                            f"{MONITORING_URL}/metrics/record",
                            json={
                                "agent_name": agent_name,
                                "task_id": task_id,
                                "timestamp": datetime.now().isoformat(),
                                "accuracy_score": 0,
                                "latency_seconds": latency,
                                "success": False,
                                "error_message": str(e),
                                "task_type": task_type,
                            },
                            timeout=2,
                        )
                    except:
                        pass

                    raise

            return sync_wrapper

    return decorator


# ============================================================================
# PYDANTIC MODELS
# ============================================================================


class TimeStudyInput(BaseModel):
    """Single time study observation"""

    process_name: str
    date: str
    observer: str
    account: str
    trial_number: str
    observed_time_seconds: float
    num_units: int
    num_workers: int
    performance_rating: float = 1.0
    notes: Optional[str] = None


class VolumeInput(BaseModel):
    """Volume forecast data"""

    process_name: str
    daily_volume: float
    unit_type: str


class ResourceInput(BaseModel):
    """Current resource allocation"""

    process_name: str
    current_workers: int
    shifts_per_day: int = 2
    hours_per_shift: float = 8.0


class SingleProcessAnalysisRequest(BaseModel):
    """Request to analyze a single process"""

    time_studies: List[TimeStudyInput]
    volume: VolumeInput
    resources: ResourceInput
    parameters: Optional[Dict] = None


class MultiProcessAnalysisRequest(BaseModel):
    """Request to analyze multiple processes"""

    processes: List[SingleProcessAnalysisRequest]


# ============================================================================
# IMPROVED EXCEL PARSER
# ============================================================================


def parse_excel_upload(file_content: bytes) -> Dict:
    """
    Parse VFP-Productivity_TMS.xlsx format

    STRUCTURE DETECTED:
    1. Process Sheets (INBOUND CHECKING, PUT-AWAY, PICKING, etc.)
       - Row 7-8: Headers with "OBSERVED TIMES (sec)" and "T1 T2 T3 ... T10"
       - Row 9+: Time study data with 10 trials per row
       - Row 13-17: Number of units (pallets, boxes, items)

    2. SUMMARY Sheet
       - Column headers: ACTIVITY, MAIN UOM, Average Volume/Day, Productivity, etc.
       - Data rows with process names and volumes

    3. B. Assmptns,Vlme,Prdctvty Sheet
       - Section "II. VOLUME" with volume data
    """
    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_content))
        sheet_names = wb.sheetnames

        print(f"\n📋 Detected {len(sheet_names)} sheets")
        print(f"   Sheets: {sheet_names[:10]}...")

        result = {"time_studies": [], "volumes": [], "resources": []}

        # ====================================================================
        # STEP 1: Parse SUMMARY sheet for VOLUMES and RESOURCES
        # ====================================================================
        if "SUMMARY" in sheet_names:
            print(f"\n📊 Parsing SUMMARY sheet for volumes and resources...")
            ws = wb["SUMMARY"]

            # Find header row (look for "ACTIVITY" and "Average Volume/Day")
            header_row_idx = None
            col_map = {}

            for row_idx in range(1, 20):
                row = [ws.cell(row_idx, col).value for col in range(1, 15)]
                row_text = [str(v).upper() if v else "" for v in row]
                row_str = " ".join(row_text)

                if "ACTIVITY" in row_str and "VOLUME" in row_str:
                    header_row_idx = row_idx

                    # Map columns
                    for col_idx, text in enumerate(row_text, 1):
                        if "ACTIVITY" in text:
                            col_map["process"] = col_idx
                        elif "MAIN UOM" in text or "UOM" in text:
                            col_map["unit"] = col_idx
                        elif "AVERAGE" in text and "VOLUME" in text:
                            col_map["volume"] = col_idx
                        elif "PRODUCTIVITY PER MANHOUR" in text:
                            col_map["productivity"] = col_idx
                        elif "MANPOWER PER SHIFT" in text:
                            col_map["workers_shift"] = col_idx
                        elif "REQUIRED MANPOWER" in text:
                            col_map["required"] = col_idx

                    print(f"   ✓ Found header at row {row_idx}")
                    print(f"   ✓ Column mapping: {col_map}")
                    break

            if header_row_idx and col_map.get("process"):
                # Parse data rows
                for row_idx in range(header_row_idx + 1, ws.max_row + 1):
                    row = [ws.cell(row_idx, col).value for col in range(1, 15)]

                    process_col = col_map.get("process", 2)
                    if not row[process_col - 1]:
                        continue

                    process_name = str(row[process_col - 1]).strip()

                    # Skip empty, totals, or separators
                    if not process_name or process_name.upper() in [
                        "",
                        "TOTAL",
                        "SUBTOTAL",
                        "GRAND TOTAL",
                    ]:
                        continue

                    # Get volume
                    volume_val = 0
                    if col_map.get("volume"):
                        vol_col = col_map["volume"]
                        try:
                            volume_val = (
                                float(row[vol_col - 1]) if row[vol_col - 1] else 0
                            )
                        except (ValueError, TypeError):
                            pass

                    # Get unit type
                    unit_type = "units"
                    if col_map.get("unit"):
                        unit_col = col_map["unit"]
                        unit_val = row[unit_col - 1]
                        if unit_val:
                            unit_type = str(unit_val).strip()

                    # Get worker count
                    workers = 2  # Default
                    if col_map.get("workers_shift"):
                        worker_col = col_map["workers_shift"]
                        try:
                            workers = (
                                int(row[worker_col - 1]) if row[worker_col - 1] else 2
                            )
                        except (ValueError, TypeError):
                            pass

                    # Add volume if found
                    if volume_val > 0:
                        result["volumes"].append(
                            {
                                "process_name": process_name,
                                "daily_volume": volume_val,
                                "unit_type": unit_type,
                            }
                        )
                        print(f"   ✓ {process_name}: {volume_val} {unit_type}/day")

                    # Add resource
                    result["resources"].append(
                        {
                            "process_name": process_name,
                            "current_workers": workers,
                            "shifts_per_day": 2,
                            "hours_per_shift": 8.0,
                        }
                    )
                    print(f"   ✓ {process_name}: {workers} workers")

        # ====================================================================
        # STEP 2: Parse PROCESS SHEETS for TIME STUDY data
        # ====================================================================

        # Detect process sheets (exclude templates, summaries, utilities)
        excluded_sheets = {
            "A. Process Flow Sample Template",
            "B. Assmptns,Vlme,Prdctvty",
            "PROD SUMMARY 2",
            "SUMMARY",
            "JACKWRAP UTILIZATION",
            "CANCEL",
            "RETURNS",
        }

        process_sheets = [s for s in sheet_names if s not in excluded_sheets]

        print(f"\n⏱️  Parsing {len(process_sheets)} process sheets for time studies...")

        for sheet_name in process_sheets:
            print(f"\n   📄 {sheet_name}")
            ws = wb[sheet_name]

            # Strategy: Look for "OBSERVED TIMES (sec)" header and T1-T10 columns
            time_study_row = None
            trial_start_col = None

            for row_idx in range(1, 15):
                row = [ws.cell(row_idx, col).value for col in range(1, 20)]
                row_text = [str(v).upper() if v else "" for v in row]
                row_str = " ".join(row_text)

                # Look for header with "OBSERVED TIMES" and trial numbers
                if "OBSERVED" in row_str and "TIMES" in row_str:
                    # Next row should have T1 T2 T3...
                    next_row = [ws.cell(row_idx + 1, col).value for col in range(1, 20)]
                    next_row_text = [str(v).upper() if v else "" for v in next_row]

                    if "T1" in next_row_text or "TRIAL" in " ".join(next_row_text):
                        time_study_row = row_idx + 2  # Data starts 2 rows after header

                        # Find where T1 starts
                        for col_idx, val in enumerate(next_row, 1):
                            if val and str(val).upper() in ["T1", "TRIAL 1", "1"]:
                                trial_start_col = col_idx
                                break

                        print(
                            f"      ✓ Found time study at row {time_study_row}, col {trial_start_col}"
                        )
                        break

            if not time_study_row or not trial_start_col:
                print(f"      ✗ No time study structure found")
                continue

            # Extract time study data
            # Typically first data row (Row 9) contains the observed times
            row = [
                ws.cell(time_study_row, col).value
                for col in range(trial_start_col, trial_start_col + 10)
            ]

            # Get element name (usually in column B or C)
            element_name = (
                ws.cell(time_study_row, 2).value
                or ws.cell(time_study_row, 3).value
                or "Process"
            )

            # Extract trial values
            trials = []
            for val in row:
                try:
                    time_val = float(val) if val else 0
                    if time_val > 0:
                        trials.append(time_val)
                except (ValueError, TypeError):
                    continue

            if not trials:
                print(f"      ✗ No valid trial data found")
                continue

            # Get number of units (look in rows 13-17 for unit counts)
            num_units = 1
            for check_row in range(time_study_row + 4, time_study_row + 10):
                val = ws.cell(check_row, trial_start_col).value
                try:
                    unit_count = int(val) if val else 0
                    if 1 <= unit_count <= 100000:
                        num_units = unit_count
                        break
                except (ValueError, TypeError):
                    continue

            # Get observer (usually in row 3-5)
            observer = "Unknown"
            for row_idx in range(3, 6):
                val = ws.cell(row_idx, 4).value or ws.cell(row_idx, 5).value
                if val and isinstance(val, str) and len(val) > 2:
                    observer = val
                    break

            # Create time study observations for each trial
            for idx, trial_time in enumerate(trials, 1):
                result["time_studies"].append(
                    {
                        "process_name": sheet_name,
                        "date": datetime.now().isoformat(),
                        "observer": observer,
                        "account": "Unknown",
                        "trial_number": f"T{idx}",
                        "observed_time_seconds": trial_time,
                        "num_units": num_units,
                        "num_workers": 1,
                        "performance_rating": 1.0,
                        "notes": f"Extracted from '{element_name}' - Trial {idx}",
                    }
                )

            print(
                f"      ✓ Extracted {len(trials)} trials, avg {np.mean(trials):.2f} sec for {num_units} unit(s)"
            )

        # ====================================================================
        # STEP 3: Fill missing data with defaults
        # ====================================================================

        # If no volumes found, create defaults from time studies
        if not result["volumes"] and result["time_studies"]:
            print(f"\n⚠️  No volumes found, creating defaults from time studies")
            processes_with_studies = set(
                ts["process_name"] for ts in result["time_studies"]
            )

            for process_name in processes_with_studies:
                # Estimate volume from average units in time study
                process_studies = [
                    ts
                    for ts in result["time_studies"]
                    if ts["process_name"] == process_name
                ]
                avg_units = np.mean([ts["num_units"] for ts in process_studies])

                result["volumes"].append(
                    {
                        "process_name": process_name,
                        "daily_volume": avg_units * 10,  # Estimate: 10 cycles per day
                        "unit_type": "units",
                    }
                )
                print(f"   ✓ {process_name}: Estimated {avg_units * 10:.0f} units/day")

        # If no resources found, create defaults
        if not result["resources"] and (result["volumes"] or result["time_studies"]):
            print(f"\n⚠️  No resources found, creating defaults")
            processes = set(v["process_name"] for v in result["volumes"]) | set(
                ts["process_name"] for ts in result["time_studies"]
            )

            for process_name in processes:
                result["resources"].append(
                    {
                        "process_name": process_name,
                        "current_workers": 2,
                        "shifts_per_day": 2,
                        "hours_per_shift": 8.0,
                    }
                )
                print(f"   ✓ {process_name}: Default 2 workers")

        wb.close()

        # Validate we got data
        if not result["time_studies"] and not result["volumes"]:
            raise HTTPException(
                status_code=400,
                detail="No workload data found. Check if Excel contains TIME STUDY data or SUMMARY sheet.",
            )

        print(f"\n✅ Parsing complete:")
        print(f"   Time studies: {len(result['time_studies'])}")
        print(f"   Volumes: {len(result['volumes'])}")
        print(f"   Resources: {len(result['resources'])}")

        return result

    except HTTPException:
        raise
    except Exception as e:
        import traceback

        traceback.print_exc()
        raise HTTPException(
            status_code=400, detail=f"Error parsing Excel file: {str(e)}"
        )


def convert_to_observations(
    time_studies: List[TimeStudyInput],
) -> List[TimeStudyObservation]:
    """Convert API input to TimeStudyObservation objects"""
    observations = []

    for ts in time_studies:
        try:
            date = datetime.fromisoformat(ts.date)
        except:
            date = datetime.now()

        obs = TimeStudyObservation(
            process_name=ts.process_name,
            observed_time_seconds=ts.observed_time_seconds,
            num_units=ts.num_units,
            num_workers=ts.num_workers,
            date=date,
            account=ts.account,
            observer=ts.observer,
        )
        observations.append(obs)

    return observations


def format_analysis_result(
    process_name: str, analysis: Dict, volume_info: Dict
) -> Dict:
    """Format analysis results for API response"""
    return {
        "process_name": process_name,
        "unit_type": volume_info.get("unit_type", "units"),
        "analysis_date": datetime.now().isoformat(),
        "productivity": {
            "per_hour": round(analysis["productivity_per_hour"], 2),
            "per_day": round(analysis["productivity_per_day"], 2),
        },
        "staffing": {
            "required_workers": analysis["required_workers"],
            "current_workers": analysis["current_workers"],
            "utilization_percent": round(analysis["utilization_percent"], 1),
            "status": analysis["status"],
        },
        "capacity": {
            "max_daily": round(analysis["max_daily_capacity"], 2),
            "throughput_hours": round(analysis["throughput_hours"], 2),
        },
        "time_study": {
            "observed_time_seconds": round(analysis["observed_time_seconds"], 2),
            "standard_time_seconds": round(analysis["standard_time_seconds"], 2),
        },
        "recommendation": analysis["recommendation"],
    }


# ============================================================================
# API ENDPOINTS (same as before, just using improved parser)
# ============================================================================


@monitor_task("workload_analysis_agent", "workload_analysis")
@app.post("/analyze/upload")
async def analyze_from_excel(file: UploadFile = File(...)):
    """
    Analyze workload from uploaded Excel file (VFP-Productivity_TMS.xlsx format)
    """
    try:
        # Validate file type
        if not file.filename.endswith((".xlsx", ".xlsm")):
            raise HTTPException(
                status_code=400, detail="File must be Excel (.xlsx or .xlsm)"
            )

        # Read file content
        content = await file.read()

        # Parse Excel using IMPROVED parser
        parsed_data = parse_excel_upload(content)

        # Group time studies by process
        time_studies_by_process = {}
        for ts in parsed_data["time_studies"]:
            process = ts["process_name"]
            if process not in time_studies_by_process:
                time_studies_by_process[process] = []
            time_studies_by_process[process].append(TimeStudyInput(**ts))

        # Create volume and resource lookups
        volumes = {v["process_name"]: v for v in parsed_data["volumes"]}
        resources = {r["process_name"]: r for r in parsed_data["resources"]}

        # Build analysis requests
        processes = []
        for process_name, time_studies in time_studies_by_process.items():
            # Get volume (or use default)
            if process_name in volumes:
                volume = volumes[process_name]
            else:
                avg_units = sum(ts.num_units for ts in time_studies) / len(time_studies)
                volume = {
                    "process_name": process_name,
                    "daily_volume": avg_units * 10,
                    "unit_type": "units",
                }
                print(
                    f"   ⚠️  {process_name}: No volume, estimated {avg_units * 10:.0f} units/day"
                )

            # Get resources (or use default)
            if process_name in resources:
                resource = resources[process_name]
            else:
                resource = {
                    "process_name": process_name,
                    "current_workers": 2,
                    "shifts_per_day": 2,
                    "hours_per_shift": 8.0,
                }
                print(
                    f"   ⚠️  {process_name}: No resource data, defaulting to 2 workers"
                )

            processes.append(
                SingleProcessAnalysisRequest(
                    time_studies=time_studies,
                    volume=VolumeInput(**volume),
                    resources=ResourceInput(**resource),
                )
            )

        if not processes:
            raise HTTPException(
                status_code=400,
                detail="No processes could be analyzed. Check if time study data is valid.",
            )

        # Analyze all processes
        analyzer = WorkloadAnalyzer()
        comparator = ProcessComparator()

        all_results = {}
        formatted_results = []

        for process_req in processes:
            observations = convert_to_observations(process_req.time_studies)

            if not observations:
                continue

            # Perform analysis
            analysis = analyzer.full_analysis(
                observations=observations,
                daily_volume=process_req.volume.daily_volume,
                current_workers=process_req.resources.current_workers,
            )

            all_results[process_req.volume.process_name] = analysis

            formatted = format_analysis_result(
                process_req.volume.process_name,
                analysis,
                {"unit_type": process_req.volume.unit_type},
            )
            formatted_results.append(formatted)

        # Identify bottleneck
        bottleneck_name, bottleneck_time = comparator.identify_bottleneck(all_results)
        system_capacity = comparator.calculate_system_capacity(all_results)

        # Rank processes by throughput
        ranked = sorted(
            all_results.items(),
            key=lambda x: x[1].get("throughput_hours", 0),
            reverse=True,
        )

        result = {
            "status": "success",
            "filename": file.filename,
            "upload_timestamp": datetime.now().isoformat(),
            "summary": {
                "total_processes": len(all_results),
                "analysis_date": datetime.now().isoformat(),
                "system_capacity": round(system_capacity, 2),
            },
            "processes": formatted_results,
            "bottleneck": {
                "process_name": bottleneck_name,
                "throughput_hours": round(bottleneck_time, 2),
                "severity": (
                    "CRITICAL"
                    if bottleneck_time > 4
                    else "WARNING" if bottleneck_time > 2 else "OK"
                ),
            },
            "ranking": [
                {
                    "rank": idx + 1,
                    "process": name,
                    "throughput_hours": round(data["throughput_hours"], 2),
                    "utilization_percent": round(data["utilization_percent"], 1),
                }
                for idx, (name, data) in enumerate(ranked)
            ],
        }

        # Cache result
        analysis_history.append(
            {
                "timestamp": datetime.now().isoformat(),
                "filename": file.filename,
                "result": result,
            }
        )

        return result

    except HTTPException:
        raise
    except Exception as e:
        import traceback

        traceback.print_exc()
        raise HTTPException(
            status_code=500, detail=f"Upload processing failed: {str(e)}"
        )


@app.get("/health")
async def health_check():
    """Health check endpoint"""
    return {
        "status": "healthy",
        "agent": "Workload Analysis Agent (IMPROVED)",
        "version": "2.0.0",
        "port": 8008,
        "parser": "VFP-Productivity_TMS.xlsx compatible",
        "features": [
            "TIME STUDY extraction from T1-T10 trials",
            "Volume from SUMMARY sheet",
            "Multi-unit support (PALLET, BOX, ITEMS, PCS)",
            "Bottleneck identification",
            "Cost-benefit analysis",
        ],
        "total_analyses": len(analysis_history),
    }


@app.get("/")
async def root():
    """Root endpoint"""
    return {
        "service": "SafexpressOps Workload Analysis Agent",
        "version": "2.0.0 (IMPROVED)",
        "endpoints": {
            "POST /analyze/upload": "Upload Excel and analyze workload",
            "GET /health": "Health check",
        },
        "documentation": "/docs",
        "excel_format": "VFP-Productivity_TMS.xlsx",
    }


# ============================================================================
# MICROSERVICE ENDPOINT (for Supervisor integration)
# ============================================================================


class ToolRequest(BaseModel):
    """Tool execution request"""

    tool: str
    inputs: Dict[str, Any]
    credentials_dict: Optional[Dict] = None


class ToolResponse(BaseModel):
    """Tool execution response"""

    success: bool
    result: Optional[Dict[str, Any]] = None
    error: Optional[str] = None


@app.post("/execute_task")
async def execute_task(request: ToolRequest):
    """Execute a workload analysis tool"""
    try:
        tool = request.tool
        inputs = request.inputs

        print(f"\n📊 Workload Agent - Tool: {tool}")

        if tool == "analyze_from_upload":
            file_path = inputs.get("file")

            if not file_path or not os.path.exists(file_path):
                return ToolResponse(success=False, error=f"File not found: {file_path}")

            # Create UploadFile object from file path
            with open(file_path, "rb") as f:
                content = f.read()

            # Create fake UploadFile
            from fastapi import UploadFile
            from io import BytesIO

            fake_file = UploadFile(
                filename=os.path.basename(file_path), file=BytesIO(content)
            )

            result = await analyze_from_excel(fake_file)

            return ToolResponse(
                success=result.get("status") == "success", result=result
            )

        else:
            return ToolResponse(success=False, error=f"Unknown tool: {tool}")

    except Exception as e:
        import traceback

        traceback.print_exc()
        return ToolResponse(success=False, error=str(e))


if __name__ == "__main__":
    import uvicorn

    uvicorn.run(app, host="0.0.0.0", port=8008)
